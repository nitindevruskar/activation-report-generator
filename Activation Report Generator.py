import re
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import traceback

# ---------------- CONFIG ----------------
GA_MTD_FILE = "Sample activation file.xlsx"
GA_MTD_SHEET_UK = "EU"
GA_MTD_SHEET_SPAIN = "USA"
GA_MTD_SHEET_GERMANY = "THA"

DB = {
    "host": "aws-1-ap-south-1.pooler.supabase.com",
    "port": 6543,
    "database": "postgres",
    "user": "postgres.jckwfzamiunzotwzpsvl",
    "password": "YOUR_SUPABASE_PASSWORD",
}

BLUE = "BDD7EE"

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ---------------- HELPERS ----------------
def find_column(df, candidates):
    cols = {re.sub(r"[\s\-_]+", "", c).upper(): c for c in df.columns}

    for cand in candidates:
        key = re.sub(r"[\s\-_]+", "", cand).upper()

        if key in cols:
            return cols[key]

    raise ValueError(f"Column not found: {candidates}")


def clean_sim(series):
    return (
        series.astype(str)
        .str.strip()
        .str.replace("'", "", regex=False)
        .str.replace('"', "", regex=False)
        .str.replace(r"\s+", "", regex=True)
        .str.lstrip("0")
        .str.lower()
    )


def fmt_date(s):
    return pd.to_datetime(
        s,
        errors="coerce"
    ).dt.strftime("%d-%m-%Y")


def add_total_row(df, label, cols):

    if not isinstance(cols, list):
        cols = [cols]

    total = {label: "Grand Total"}

    for c in cols:
        total[c] = pd.to_numeric(
            df[c],
            errors="coerce"
        ).sum()

    return pd.concat(
        [df, pd.DataFrame([total])],
        ignore_index=True
    )


def format_table(ws, start_row, start_col, df):

    rows = df.shape[0] + 1
    cols = df.shape[1]

    fill = PatternFill(
        start_color=BLUE,
        end_color=BLUE,
        fill_type="solid"
    )

    bold = Font(bold=True)

    # HEADER
    for c in range(start_col, start_col + cols):

        cell = ws.cell(start_row, c)

        cell.fill = fill
        cell.font = bold
        cell.border = THIN

    # DATA
    for r in range(start_row + 1, start_row + rows):

        for c in range(start_col, start_col + cols):

            cell = ws.cell(r, c)

            cell.border = THIN

            if ws.cell(r, start_col).value == "Grand Total":

                cell.fill = fill
                cell.font = bold


def autofit_and_center(ws):

    for col in ws.columns:

        max_len = 0

        letter = col[0].column_letter

        for cell in col:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if cell.value:
                max_len = max(
                    max_len,
                    len(str(cell.value))
                )

        ws.column_dimensions[letter].width = max_len + 2


def resolve_country(sim, sim_map):

    if sim in sim_map:
        return sim_map[sim]

    sim_minus_last = sim[:-1]

    if sim_minus_last in sim_map:
        return sim_map[sim_minus_last]

    return "Unknown"


# ==========================================
# ONLY FIXED FUNCTION
# ==========================================
def build_country_summary(df, fixed_order):

    # CLEAN COUNTRY VALUES
    df["country"] = (
        df["country"]
        .fillna("Unknown")
        .astype(str)
        .str.strip()
    )

    # REMOVE DUPLICATE COUNTRY NAMES
    fixed_order = list(dict.fromkeys(fixed_order))

    pivot = df.pivot_table(
        index="ga_date",
        columns="country",
        values="sim_no",
        aggfunc="count",
        fill_value=0
    )

    # REMOVE DUPLICATE COLUMNS
    pivot = pivot.loc[:, ~pivot.columns.duplicated()]

    for c in fixed_order:
        if c not in pivot.columns:
            pivot[c] = 0

    cols = fixed_order.copy()

    if "Unknown" in pivot.columns:
        cols.append("Unknown")

    # REMOVE DUPLICATES AGAIN
    cols = list(dict.fromkeys(cols))

    pivot = pivot[cols]

    for c in cols:
        pivot[c] = pivot[c].replace(0, "")

    pivot["Grand Total"] = (
        pivot.replace("", 0)
        .astype(int)
        .sum(axis=1)
    )

    pivot = (
        pivot.reset_index()
        .rename(columns={"ga_date": "Date"})
    )

    # REMOVE DUPLICATE COLUMNS AFTER RESET INDEX
    pivot = pivot.loc[:, ~pivot.columns.duplicated()]

    total = {"Date": "Grand Total"}

    for c in cols:
        total[c] = (
            pivot[c]
            .replace("", 0)
            .astype(int)
            .sum()
        )

    total["Grand Total"] = pivot["Grand Total"].sum()

    total_df = pd.DataFrame([total])

    # FINAL SAFE CONCAT
    final_df = pd.concat(
        [
            pivot.reset_index(drop=True),
            total_df.reset_index(drop=True)
        ],
        ignore_index=True
    )

    return final_df


# ---------------- MAIN ----------------
def main():

    try:

        print("🚀 Generating Activation Report......")

        engine = create_engine(
            f"postgresql+psycopg2://"
            f"{DB['user']}:{DB['password']}"
            f"@{DB['host']}:{DB['port']}"
            f"/{DB['database']}"
        )

        # ================= UK =================
        df_uk = pd.read_excel(
            GA_MTD_FILE,
            GA_MTD_SHEET_UK,
            dtype=str
        )

        sim_col = find_column(
            df_uk,
            ["ICCID", "SIM_NO"]
        )

        date_col = find_column(
            df_uk,
            ["GA_DATE", "DATE"]
        )

        df_uk[sim_col] = clean_sim(
            df_uk[sim_col]
        )

        df_uk[date_col] = pd.to_datetime(
            df_uk[date_col],
            errors="coerce"
        )

        last_date = df_uk[date_col].max()

        cutoff_day = last_date.day

        file_name = (
            f"Lebara Activation Details "
            f"{last_date.strftime('%d.%m.%Y')}.xlsx"
        )

        sim_master_uk = pd.read_sql(
            """
            SELECT sim_no, country, group_name
            FROM public.sim_master_europe_tsim
            """,
            engine
        )

        sim_master_uk["sim_no"] = clean_sim(
            sim_master_uk["sim_no"]
        )

        sim_master_uk["group_name"] = (
            sim_master_uk["group_name"]
            .astype(str)
            .str.upper()
            .str.replace(
                r"[\s\-_]+",
                "",
                regex=True
            )
        )

        merged = pd.merge(
            df_uk,
            sim_master_uk,
            left_on=sim_col,
            right_on="sim_no",
            how="left"
        )

        uk_non_vcas = merged[
            merged["group_name"] != "TLSVCAS"
        ]

        uk_vcas = merged[
            merged["group_name"] == "TLSVCAS"
        ]

        df_uk_out = (
            uk_non_vcas[
                [sim_col, date_col, "country"]
            ]
            .rename(
                columns={
                    sim_col: "sim_no",
                    date_col: "ga_date"
                }
            )
        )

        df_uk_out["ga_date"] = fmt_date(
            df_uk_out["ga_date"]
        )

        df_tls_out = (
            uk_vcas[
                [sim_col, date_col, "country"]
            ]
            .rename(
                columns={
                    sim_col: "sim_no",
                    date_col: "ga_date"
                }
            )
        )

        df_tls_out["ga_date"] = fmt_date(
            df_tls_out["ga_date"]
        )

        # ================= UK COMPARISON =================
        this_year = last_date.year
        this_month = last_date.month

        this_hdr = (
            f"{last_date.strftime('%b-%y')} "
            f"(1-{cutoff_day})"
        )

        last_hdr = (
            f"{(last_date.replace(year=this_year-1)).strftime('%b-%y')} "
            f"(1-{cutoff_day})"
        )

        this_counts = (
            uk_non_vcas
            .groupby("country")
            .size()
            .reset_index(name=this_hdr)
        )

        last_df_sql = pd.read_sql(
            f"""
            SELECT sim_no, activation_date
            FROM public.activation_europe_tsim
            WHERE date_part('year', activation_date)={this_year-1}
            AND date_part('month', activation_date)={this_month}
            AND date_part('day', activation_date) <= {cutoff_day}
            """,
            engine
        )

        last_df_sql["sim_no"] = clean_sim(
            last_df_sql["sim_no"]
        )

        last_df_sql = last_df_sql.merge(
            sim_master_uk,
            on="sim_no",
            how="left"
        )

        last_df_sql = last_df_sql[
            last_df_sql["group_name"] != "TLSVCAS"
        ]

        last_counts = (
            last_df_sql
            .groupby("country")
            .size()
            .reset_index(name=last_hdr)
        )

        uk_comp = pd.merge(
            this_counts,
            last_counts,
            on="country",
            how="outer"
        ).fillna(0)

        uk_comp["Gap"] = (
            uk_comp[this_hdr]
            - uk_comp[last_hdr]
        )

        uk_comp["% Change"] = (
            (
                uk_comp["Gap"]
                / uk_comp[last_hdr].replace(0, pd.NA)
            ) * 100
        ).fillna(0).round(2)

        uk_comp = uk_comp.sort_values(
            by="Gap",
            ascending=True
        )

        uk_comp = add_total_row(
            uk_comp,
            "country",
            [this_hdr, last_hdr, "Gap"]
        )

        ti = uk_comp.index[-1]

        tg = uk_comp.loc[ti, "Gap"]
        tl = uk_comp.loc[ti, last_hdr]

        uk_comp.at[ti, "% Change"] = (
            round((tg / tl) * 100, 2)
            if tl != 0 else 0
        )

        uk_comp.rename(
            columns={"country": "Country"},
            inplace=True
        )

        # ================= SUMMARIES =================
        summary_uk = add_total_row(
            df_uk.groupby(date_col)
            .size()
            .reset_index(name="Count")
            .rename(columns={date_col: "Date"}),
            "Date",
            "Count"
        )

        summary_uk["Date"] = fmt_date(
            summary_uk["Date"]
        )

        summary_uk.loc[
            summary_uk["Date"].isna(),
            "Date"
        ] = "Grand Total"

        summary_tls = add_total_row(
            df_tls_out.groupby("ga_date")
            .size()
            .reset_index(name="Count")
            .rename(columns={"ga_date": "Date"}),
            "Date",
            "Count"
        )

        # ================= YOY =================
        this_yoy_hdr = datetime(
            this_year,
            this_month,
            1
        ).strftime("%b-%y")

        last_yoy_hdr = datetime(
            this_year - 1,
            this_month,
            1
        ).strftime("%b-%y")

        this_daily = uk_non_vcas.copy()

        this_daily["Day"] = (
            this_daily[date_col].dt.day
        )

        this_daily = (
            this_daily.groupby("Day")
            .size()
            .reset_index(name=this_yoy_hdr)
        )

        last_daily = last_df_sql.copy()

        last_daily["Day"] = (
            pd.to_datetime(
                last_daily["activation_date"]
            ).dt.day
        )

        last_daily = (
            last_daily.groupby("Day")
            .size()
            .reset_index(name=last_yoy_hdr)
        )

        yoy = pd.merge(
            this_daily[
                this_daily["Day"] <= cutoff_day
            ],
            last_daily[
                last_daily["Day"] <= cutoff_day
            ],
            on="Day",
            how="outer"
        ).fillna(0)

        yoy["Diff"] = (
            yoy[this_yoy_hdr]
            - yoy[last_yoy_hdr]
        )

        yoy = add_total_row(
            yoy,
            "Day",
            [this_yoy_hdr, last_yoy_hdr, "Diff"]
        )

        # ================= SPAIN & GERMANY =================
        def process_ext(sheet, master_tbl, act_tbl):

            df = pd.read_excel(
                GA_MTD_FILE,
                sheet,
                dtype=str
            )

            sc = find_column(
                df,
                ["ICCID", "SIM_NO"]
            )

            df[sc] = clean_sim(df[sc])

            m = pd.read_sql(
                f"""
                SELECT sim_no, country
                FROM {master_tbl}
                """,
                engine
            )

            m["sim_no"] = clean_sim(
                m["sim_no"]
            )

            s_map = dict(
                zip(m["sim_no"], m["country"])
            )

            a = pd.read_sql(
                f"""
                SELECT sim_no, activation_date
                FROM {act_tbl}
                """,
                engine
            )

            a["sim_no"] = clean_sim(
                a["sim_no"]
            )

            df = pd.merge(
                df,
                a,
                left_on=sc,
                right_on="sim_no",
                how="left"
            )

            df["country"] = df[sc].apply(
                lambda x: resolve_country(x, s_map)
            )

            df["ga_date"] = fmt_date(
                df["activation_date"]
            )

            return df[
                ["sim_no", "ga_date", "country"]
            ]

        df_sp = process_ext(
            GA_MTD_SHEET_SPAIN,
            "public.sim_master_usa_tsim",
            "public.activation_usa_tsim"
        )

        df_de = process_ext(
            GA_MTD_SHEET_GERMANY,
            "public.sim_master_thailand_ais",
            "public.activation_thailand_ais"
        )

        spain_summary = build_country_summary(
            df_sp,
            [
                "Morocco",
                "Peru",
                "Ecuador",
                "Ghana",
                "USA",
                "China",
                "Colombia",
                "Mexico",
                "EU",
                "Turkey",
                "India",
                "Senegal",
                "Argentina",
		"Equatorial Guinea",
		"South Africa",
		"Philippines",
		"Bolivia",
		"Chile",
		"South Africa"
            ]
        )

        germany_summary = build_country_summary(
            df_de,
            sorted(df_de["country"].unique().tolist())
        )

        # ================= WRITE =================
        with pd.ExcelWriter(
            file_name,
            engine="openpyxl"
        ) as writer:

            df_uk_out.to_excel(
                writer,
                "EU",
                index=False
            )

            uk_comp.to_excel(
                writer,
                "Europe country wise comparison",
                index=False
            )

            df_tls_out.to_excel(
                writer,
                "TLS_VCAS",
                index=False
            )

            df_sp.to_excel(
                writer,
                "USA",
                index=False
            )

            df_de.to_excel(
                writer,
                "THA",
                index=False
            )

            summary_uk.to_excel(
                writer,
                "Summary",
                startrow=1,
                startcol=0,
                index=False
            )

            yoy.to_excel(
                writer,
                "Summary",
                startrow=1,
                startcol=summary_uk.shape[1] + 2,
                index=False
            )

            summary_tls.to_excel(
                writer,
                "Summary",
                startrow=1,
                startcol=summary_uk.shape[1] + yoy.shape[1] + 4,
                index=False
            )

            spain_summary.to_excel(
                writer,
                "Summary",
                startrow=1,
                startcol=summary_uk.shape[1] + yoy.shape[1] + summary_tls.shape[1] + 6,
                index=False
            )

            germany_summary.to_excel(
                writer,
                "Summary",
                startrow=1,
                startcol=summary_uk.shape[1] + yoy.shape[1] + summary_tls.shape[1] + spain_summary.shape[1] + 8,
                index=False
            )

        # ================= FORMATTING =================
        wb = load_workbook(file_name)

        for s, d in [
            ("EU", df_uk_out),
            ("Europe country wise comparison", uk_comp),
            ("TLS_VCAS", df_tls_out),
            ("USA", df_sp),
            ("THA", df_de)
        ]:

            ws = wb[s]

            format_table(ws, 1, 1, d)

            autofit_and_center(ws)

        ws_sum = wb["Summary"]

        format_table(ws_sum, 2, 1, summary_uk)

        format_table(
            ws_sum,
            2,
            summary_uk.shape[1] + 3,
            yoy
        )

        format_table(
            ws_sum,
            2,
            summary_uk.shape[1] + yoy.shape[1] + 5,
            summary_tls
        )

        format_table(
            ws_sum,
            2,
            summary_uk.shape[1] + yoy.shape[1] + summary_tls.shape[1] + 7,
            spain_summary
        )

        format_table(
            ws_sum,
            2,
            summary_uk.shape[1] + yoy.shape[1] + summary_tls.shape[1] + spain_summary.shape[1] + 9,
            germany_summary
        )

        autofit_and_center(ws_sum)

        for row in ws_sum.iter_rows(
            min_row=2,
            max_row=2 + summary_uk.shape[0],
            min_col=1,
            max_col=summary_uk.shape[1]
        ):

            if row[0].value == "Grand Total":

                for cell in row:

                    cell.fill = PatternFill(
                        start_color=BLUE,
                        end_color=BLUE,
                        fill_type="solid"
                    )

                    cell.font = Font(bold=True)

                    cell.border = THIN

        wb.save(file_name)

        print("✅ Sorted report generated without arrows")

    except Exception:
        traceback.print_exc()


if __name__ == "__main__":
    main()
